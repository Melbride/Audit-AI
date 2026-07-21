"""
Financial statement starter template.

Generates a blank multi-sheet Excel workbook for the client/accountant to
fill in at the beginning of engagement work: a Trial Balance sheet
(matching the GL upload column structure) plus a standard 3-statement
skeleton (Balance Sheet, Income Statement, Cash Flow Statement).

Requires: pip install openpyxl --break-system-packages

Wire this into main.py with:

    from statement_template import router as statement_template_router
    app.include_router(statement_template_router)
"""

import io

from fastapi import APIRouter, HTTPException, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from database import get_db

router = APIRouter(prefix="/api/engagements", tags=["statement-template"])

HEADER_FILL = PatternFill(start_color="1E3A5F", end_color="1E3A5F", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF")
SECTION_FONT = Font(bold=True, size=12)
TOTAL_FONT = Font(bold=True)


def _style_header_row(ws, row_num: int, col_count: int):
    for col in range(1, col_count + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="left")


def _autosize(ws, widths: dict):
    for col_letter, width in widths.items():
        ws.column_dimensions[col_letter].width = width


def _build_trial_balance_sheet(wb: Workbook, meta: dict):
    ws = wb.active
    ws.title = "Trial Balance"

    ws["A1"] = f"{meta['company_name']} — Trial Balance"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Engagement: {meta['engagement_name']}  |  Period: {meta['period_start']} to {meta['period_end']}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    headers = ["AccountNumber", "AccountName", "Debit", "Credit", "Dept", "CostCenter", "Description", "Currency"]
    header_row = 4
    for i, h in enumerate(headers, start=1):
        ws.cell(row=header_row, column=i, value=h)
    _style_header_row(ws, header_row, len(headers))

    # A handful of blank rows ready for data entry
    for r in range(header_row + 1, header_row + 21):
        ws.cell(row=r, column=1)

    total_row = header_row + 21
    ws.cell(row=total_row, column=2, value="TOTAL").font = TOTAL_FONT
    ws.cell(row=total_row, column=3, value=f"=SUM(C{header_row + 1}:C{total_row - 1})").font = TOTAL_FONT
    ws.cell(row=total_row, column=4, value=f"=SUM(D{header_row + 1}:D{total_row - 1})").font = TOTAL_FONT

    _autosize(ws, {"A": 16, "B": 28, "C": 14, "D": 14, "E": 10, "F": 12, "G": 30, "H": 10})


def _build_balance_sheet(wb: Workbook, meta: dict):
    ws = wb.create_sheet("Balance Sheet")
    ws["A1"] = f"{meta['company_name']} — Balance Sheet"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"As at: {meta['period_end']}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    row = 4
    sections = [
        ("ASSETS", None),
        ("Current Assets", ["Cash and Cash Equivalents", "Accounts Receivable", "Inventory", "Prepaid Expenses"]),
        ("Non-Current Assets", ["Property, Plant & Equipment", "Intangible Assets", "Long-term Investments"]),
        ("TOTAL ASSETS", "total"),
        (None, None),
        ("LIABILITIES", None),
        ("Current Liabilities", ["Accounts Payable", "Accrued Expenses", "Short-term Debt"]),
        ("Non-Current Liabilities", ["Long-term Debt", "Deferred Tax Liabilities"]),
        ("TOTAL LIABILITIES", "total"),
        (None, None),
        ("EQUITY", None),
        ("", ["Share Capital", "Retained Earnings"]),
        ("TOTAL EQUITY", "total"),
        (None, None),
        ("TOTAL LIABILITIES AND EQUITY", "total"),
    ]
    for label, items in sections:
        if label is None:
            row += 1
            continue
        if items == "total":
            ws.cell(row=row, column=1, value=label).font = TOTAL_FONT
            ws.cell(row=row, column=2, value=0)
            row += 2
            continue
        if label:
            ws.cell(row=row, column=1, value=label).font = SECTION_FONT
            row += 1
        for item in (items or []):
            ws.cell(row=row, column=1, value=f"    {item}")
            ws.cell(row=row, column=2, value=0)
            row += 1

    _autosize(ws, {"A": 34, "B": 16})


def _build_income_statement(wb: Workbook, meta: dict):
    ws = wb.create_sheet("Income Statement")
    ws["A1"] = f"{meta['company_name']} — Income Statement"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Period: {meta['period_start']} to {meta['period_end']}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    row = 4
    lines = [
        ("Revenue", None),
        ("Cost of Sales", None),
        ("GROSS PROFIT", "total"),
        (None, None),
        ("Operating Expenses", ["Salaries and Wages", "Rent", "Utilities", "Depreciation", "Other Operating Expenses"]),
        ("Total Operating Expenses", "total"),
        (None, None),
        ("OPERATING INCOME", "total"),
        ("Other Income / (Expenses)", None),
        ("Tax Expense", None),
        ("NET INCOME", "total"),
    ]
    for label, items in lines:
        if label is None:
            row += 1
            continue
        if items == "total":
            ws.cell(row=row, column=1, value=label).font = TOTAL_FONT
            ws.cell(row=row, column=2, value=0)
            row += 2
            continue
        if isinstance(items, list):
            ws.cell(row=row, column=1, value=label).font = SECTION_FONT
            row += 1
            for item in items:
                ws.cell(row=row, column=1, value=f"    {item}")
                ws.cell(row=row, column=2, value=0)
                row += 1
        else:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=0)
            row += 1

    _autosize(ws, {"A": 34, "B": 16})


def _build_cash_flow_statement(wb: Workbook, meta: dict):
    ws = wb.create_sheet("Cash Flow Statement")
    ws["A1"] = f"{meta['company_name']} — Cash Flow Statement"
    ws["A1"].font = Font(bold=True, size=14)
    ws["A2"] = f"Period: {meta['period_start']} to {meta['period_end']}"
    ws["A2"].font = Font(italic=True, size=10, color="666666")

    row = 4
    sections = [
        ("Operating Activities", ["Net Income", "Depreciation & Amortization", "Changes in Working Capital"]),
        ("Net Cash from Operating Activities", "total"),
        (None, None),
        ("Investing Activities", ["Purchase of Property/Equipment", "Proceeds from Asset Sales"]),
        ("Net Cash from Investing Activities", "total"),
        (None, None),
        ("Financing Activities", ["Proceeds from Debt", "Repayment of Debt", "Dividends Paid"]),
        ("Net Cash from Financing Activities", "total"),
        (None, None),
        ("NET CHANGE IN CASH", "total"),
        ("Cash at Beginning of Period", None),
        ("Cash at End of Period", "total"),
    ]
    for label, items in sections:
        if label is None:
            row += 1
            continue
        if items == "total":
            ws.cell(row=row, column=1, value=label).font = TOTAL_FONT
            ws.cell(row=row, column=2, value=0)
            row += 2
            continue
        if isinstance(items, list):
            ws.cell(row=row, column=1, value=label).font = SECTION_FONT
            row += 1
            for item in items:
                ws.cell(row=row, column=1, value=f"    {item}")
                ws.cell(row=row, column=2, value=0)
                row += 1
        else:
            ws.cell(row=row, column=1, value=label)
            ws.cell(row=row, column=2, value=0)
            row += 1

    _autosize(ws, {"A": 34, "B": 16})


@router.get("/{engagement_id}/statement-template")
def download_statement_template(engagement_id: int, db=Depends(get_db)):
    """
    Generate a starter Excel workbook for this engagement: a Trial Balance
    sheet (matching the GL upload structure) plus a Balance Sheet, Income
    Statement, and Cash Flow Statement skeleton — for the client/accountant
    to fill in at the start of the work.
    """
    cursor = db.cursor(dictionary=True)
    cursor.execute(
        """
        SELECT e.engagement_id, e.engagement_name, e.start_date, e.end_date, c.company_name
        FROM engagements e
        LEFT JOIN clients c ON e.client_id = c.client_id
        WHERE e.engagement_id = %s
        """,
        (engagement_id,),
    )
    engagement = cursor.fetchone()
    if not engagement:
        raise HTTPException(status_code=404, detail="Engagement not found")

    meta = {
        "company_name": engagement.get("company_name") or "Client",
        "engagement_name": engagement["engagement_name"],
        "period_start": str(engagement.get("start_date") or ""),
        "period_end": str(engagement.get("end_date") or ""),
    }

    wb = Workbook()
    _build_trial_balance_sheet(wb, meta)
    _build_balance_sheet(wb, meta)
    _build_income_statement(wb, meta)
    _build_cash_flow_statement(wb, meta)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    safe_name = "".join(c if c.isalnum() or c in "-_" else "_" for c in meta["engagement_name"])
    filename = f"financial_statements_template_{safe_name}.xlsx"

    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )