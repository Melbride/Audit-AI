from openpyxl import load_workbook

ROW_ID_COLUMN = "_row_id"
RENAME_MATCH_THRESHOLD = 0.9

# Compare two cell values for a real difference, tolerant of numeric formatting
def values_differ(original_value, uploaded_value) -> bool:
    """
    Compare two cell values for a real difference, tolerant of numeric formatting
    differences (e.g. 45000000.0 vs 45000000) that aren't actual edits made by the auditor.
    """
    orig_str = "" if original_value is None else str(original_value).strip()
    up_str = "" if uploaded_value is None else str(uploaded_value).strip()
    if orig_str == up_str:
        return False

    # Try numeric comparison so formatting differences aren't treated as real edits
    try:
        if float(orig_str) == float(up_str):
            return False
    except (ValueError, TypeError):
        pass

    return True

# Find the changes between a snapshot and a re-uploaded file.
def _match_score(snapshot_rows_by_id, uploaded_rows_by_id, shared_row_ids, old_name, new_name):
    """
    Compares one candidate old column against one candidate new column, across every
    row present in both the snapshot and the upload. Returns (score, compared_count).
    score is the fraction of compared values that are identical (0.0 to 1.0).
    compared_count is how many rows actually contributed evidence — rows where BOTH
    sides are blank are skipped entirely, since two unrelated empty columns would
    otherwise "match" with no real evidence behind it. compared_count == 0 means this
    candidate provided no information at all and must never be treated as a winner,no matter how the score divides out.
    """
    compared = 0
    matched = 0
    for row_id in shared_row_ids:
        old_value = snapshot_rows_by_id[row_id].get(old_name)
        new_value = uploaded_rows_by_id[row_id].get(new_name)
        old_is_blank = old_value is None or str(old_value).strip() == ""
        new_is_blank = new_value is None or str(new_value).strip() == ""
        if old_is_blank and new_is_blank:
            continue
        compared += 1
        if not values_differ(old_value, new_value):
            matched += 1
    if compared == 0:
        return 0.0, 0
    return matched / compared, compared

# Resolve any ambiguities when a column is both renamed and deleted.
def _resolve_renames(missing_names, added_names, snapshot_rows_by_id, uploaded_rows_by_id):
    """
    Figures out which missing column name(s) each added column name actually replaces,
    for ANY number of simultaneous renames in one upload — not just a single rename.

    Approach: build a score for every (new_name, old_name) candidate pairing using the
    underlying cell data (see _match_score), then greedily assign the single best-scoring
    pairing first, remove both names from the pool, and repeat. This is the standard way
    to safely resolve a many-to-many matching problem without needing anything more
    elaborate — at each step we only commit to the pairing we're most confident about,
    so an uncertain pairing never gets forced into matching just because it's the only
    one left.

    A pairing is only ever accepted if its score clears RENAME_MATCH_THRESHOLD (90%) AND
    it actually has comparable data (compared > 0). Any new_name that can't be confidently
    paired with a missing_name this way is left unresolved — the caller treats that
    specific new_name as an unexplained addition (ambiguous), while every OTHER pairing
    that WAS confidently resolved still goes through normally. This means one unclear
    rename among several no longer blocks the rest, unlike the single-rename version
    where any uncertainty failed the whole upload.
    Returns (renamed_columns, unresolved_new_names) where renamed_columns is
    {old_name: new_name} for every confidently-resolved pairing, and unresolved_new_names
    is the set of added names that could not be confidently paired with anything.
    """
    if not added_names:
        return {}, set()

    shared_row_ids = set(snapshot_rows_by_id.keys()) & set(uploaded_rows_by_id.keys())

    # Build every candidate pairing's score up front.
    candidates = []  # list of (score, compared_count, new_name, old_name)
    for new_name in added_names:
        for old_name in missing_names:
            score, compared = _match_score(
                snapshot_rows_by_id, uploaded_rows_by_id, shared_row_ids, old_name, new_name
            )
            if compared > 0 and score >= RENAME_MATCH_THRESHOLD:
                candidates.append((score, compared, new_name, old_name))

    # Greedily assign the strongest pairing first, then the next-strongest among what's
    # left, and so on. A pairing is only withheld as a genuine tie when something ELSE
    candidates.sort(key=lambda c: (-c[0], -c[1]))
    renamed_columns = {}
    claimed_new_names = set()
    claimed_old_names = set()
    blocked_new_names = set()
    remaining = [c for c in candidates]

    while remaining:
        # Drop anything already claimed or already blocked by an earlier tie.
        remaining = [
            c for c in remaining
            if c[2] not in claimed_new_names and c[3] not in claimed_old_names
            and c[2] not in blocked_new_names
        ]
        if not remaining:
            break

        top_score, top_compared = remaining[0][0], remaining[0][1]
        tied_at_top = [c for c in remaining if c[0] == top_score and c[1] == top_compared]

        # Among the candidates tied at this score level, find genuine conflicts: more
        # than one candidate wanting the same new_name, or more than one wanting the same old_name.
        new_name_counts = {}
        old_name_counts = {}
        for score, compared, new_name, old_name in tied_at_top:
            new_name_counts[new_name] = new_name_counts.get(new_name, 0) + 1
            old_name_counts[old_name] = old_name_counts.get(old_name, 0) + 1
        contested_new_names = {n for n, count in new_name_counts.items() if count > 1}
        contested_old_names = {n for n, count in old_name_counts.items() if count > 1}

        # If a new_name is contested (more than one old_name wants it) or an old_name is contested
        for score, compared, new_name, old_name in tied_at_top:
            if new_name in contested_new_names or old_name in contested_old_names:
                # Genuine, unresolvable conflict at this score level — withhold this
                # new_name entirely rather than guess which old_name it really belongs to.
                blocked_new_names.add(new_name)
            else:
                renamed_columns[old_name] = new_name
                claimed_new_names.add(new_name)
                claimed_old_names.add(old_name)

        remaining = [c for c in remaining if c[0] != top_score or c[1] != top_compared]
    unresolved_new_names = added_names - claimed_new_names
    return renamed_columns, unresolved_new_names

# Compare a re-uploaded file against the snapshot of what was originally downloaded.
def diff_uploaded_against_snapshot(uploaded_file_path: str, snapshot_rows: list) -> dict:
    """
    Read a re-uploaded, auditor-edited Excel file (from the cleaning workbook export) and
    compare it against the saved snapshot of what was originally downloaded.

    Rows are matched by the hidden "_row_id" column written into every export, NOT by
    position. This means the auditor can freely delete rows (e.g. a row they judged to be
    unreliable or unnecessary) and the system still correctly identifies exactly which
    original row is missing, rather than misaligning the comparison or rejecting the file
    outright just because the row count changed.

    COLUMN RENAME RULE: any number of simultaneous column renames are supported in one
    upload. Each new column name's data is compared against every missing column's data,
    and pairings are assigned greedily by strongest match first (see _resolve_renames) —
    this lets an auditor rename several columns AND delete several others in a single
    editing session without needing multiple upload rounds, as long as each rename's
    underlying data is clearly distinguishable from every other candidate.

    A pairing is only trusted when it clears a 90% value-match threshold AND has actual
    comparable data (not all blank). Any new column name that can't be confidently traced
    to exactly one missing column this way is reported as an "unresolved" addition rather
    than guessed — but, importantly, any OTHER rename in the same upload that WAS resolved
    confidently still goes through normally. Only the genuinely unclear renames block;
    everything else proceeds.

    Renames are deliberately NOT inferred from column position, for the same reason as
    before: schema drift between the snapshot and the live mapping (e.g. a derived column
    like tax_amount appearing only after a different column is resolved) can shift column
    order in ways unrelated to anything the auditor actually did, and position-based
    pairing previously misattributed renames to completely untouched columns because of it.

    Returns a dict with:
      "corrections": list of {row_index, column, original_value, corrected_value} for every
        data cell that actually changed, keyed by the ORIGINAL (snapshot) column name.
      "deleted_row_ids": list of original row indices present in the snapshot but missing
        from the uploaded file — rows the auditor removed.
      "deleted_columns": list of column names present in the snapshot but missing from the
        uploaded file, where no rename was recognized — i.e. genuine deletions only.
      "renamed_columns": dict of {old_name: new_name} for every confidently-resolved
        rename this round. The caller should apply each of these directly to the column
        mapping (set mapped_to to the new name) regardless of whether the column was
        previously unknown or already mapped.
      "unresolved_columns": list of new column names that appeared in the upload but
        could not be confidently traced back to any missing column. These are NOT renames
        and NOT plain deletions — they're genuinely new, unexplained columns (or renames
        too ambiguous to trust). The caller should surface these specifically rather than
        silently accepting or rejecting them.
      "ambiguous_changes": True only when there is at least one unresolved_columns entry,
        kept for backward compatibility with callers checking this single flag. Prefer
        checking unresolved_columns directly for which specific column needs attention.
      "new_row_ids": list of row ids found in the uploaded file that were never in the
        snapshot — this should not normally happen since the hidden id column is locked
        from being repurposed, but is reported for safety.
    """
    wb = load_workbook(uploaded_file_path, data_only=True)
    if "Cleaned Data" not in wb.sheetnames:
        raise ValueError("Uploaded file does not contain a 'Cleaned Data' sheet. Please upload the file exactly as it was downloaded.")
    ws = wb["Cleaned Data"]

    raw_headers = [cell.value for cell in ws[1]]
    if not raw_headers or raw_headers[0] != ROW_ID_COLUMN:
        raise ValueError(
            "Uploaded file is missing its row tracking column. Please upload the file "
            "exactly as downloaded, without removing or modifying the hidden first column."
        )

    # Strip any remaining [UNRESOLVED] prefix to get each column's real underlying name.
    clean_headers = []
    for h in raw_headers[1:]:
        if h == "Issues":
            continue
        if h and str(h).startswith("[UNRESOLVED]"):
            clean_headers.append(str(h).replace("[UNRESOLVED]", "", 1).strip())
        else:
            clean_headers.append(h)

    # Reject immediately if the uploaded file has two columns with the same header
    seen_headers = {}
    duplicate_headers = set()
    for h in clean_headers:
        if h in seen_headers:
            duplicate_headers.add(h)
        seen_headers[h] = True
    if duplicate_headers:
        names = ", ".join(f"'{h}'" for h in sorted(duplicate_headers))
        raise ValueError(
            f"The uploaded file has more than one column named {names}. This usually "
            f"happens when a column was renamed to a name that another column already "
            f"uses. Please give each column a unique name and upload again."
        )

    # Build a lookup of uploaded rows keyed by their row id, reading every data row
    uploaded_rows_by_id = {}
    for row_cells in ws.iter_rows(min_row=2, values_only=False):
        row_id_value = row_cells[0].value
        if row_id_value is None:
            continue  # skip any fully blank trailing row
        row_id = int(row_id_value)
        row_data = {}
        for col_offset, col_name in enumerate(clean_headers, start=1):
            row_data[col_name] = row_cells[col_offset].value
        uploaded_rows_by_id[row_id] = row_data

    # Close the workbook now that every value needed has been read out of it. On Windows,
    # openpyxl keeps the underlying file handle open until close() is called explicitly —
    # without this, the caller's attempt to delete the temp upload file right after this
    # function returns fails with PermissionError: file in use by another process.
    wb.close()

    # Snapshot columns (excluding the internal _row_index key)
    snapshot_columns = set()
    snapshot_rows_by_id = {}
    for snap_row in snapshot_rows:
        row_id = snap_row["_row_index"]
        snapshot_rows_by_id[row_id] = snap_row
        snapshot_columns.update(k for k in snap_row.keys() if k != "_row_index")

    uploaded_columns = set(clean_headers)

    # Names missing from the upload and names newly present in the upload, by SET.
    missing_names = snapshot_columns - uploaded_columns
    added_names = uploaded_columns - snapshot_columns

    # Resolve ALL renames at once (zero, one, or several), using value-matching to pair
    # each added name with the missing name its data actually belongs to. Any added name
    # that can't be confidently paired ends up in unresolved_columns — and crucially,
    # other renames in the same batch that WERE resolved confidently are unaffected by
    # that one uncertain case.
    renamed_columns, unresolved_columns = _resolve_renames(
        missing_names, added_names, snapshot_rows_by_id, uploaded_rows_by_id
    )

    # Genuine deletions: missing names that were not claimed as the source of any rename.
    deleted_columns = sorted(missing_names - set(renamed_columns.keys()))

    ambiguous_changes = len(unresolved_columns) > 0

    # Rows present in the snapshot but missing from the upload entirely = deliberately deleted rows
    deleted_row_ids = sorted(set(snapshot_rows_by_id.keys()) - set(uploaded_rows_by_id.keys()))

    # Rows present in the upload that were never in the snapshot — should not normally
    # happen since the id column is locked, reported for safety/debugging
    new_row_ids = sorted(set(uploaded_rows_by_id.keys()) - set(snapshot_rows_by_id.keys()))

    # Diff cell values only for rows that exist in both. For columns, use the snapshot's
    # name for any column unchanged or recognized as a rename, so corrections key against
    # the name the mapping/dataframe still uses today (pre-rename), not the new label.
    # Unresolved new columns are deliberately excluded here — we don't know what original
    # column (if any) they correspond to, so there is nothing safe to diff them against.
    diffable_columns = {}  # snapshot_name -> uploaded_name
    for col_name in snapshot_columns & uploaded_columns:
        diffable_columns[col_name] = col_name
    for old_name, new_name in renamed_columns.items():
        diffable_columns[old_name] = new_name

    # Build a list of corrections to apply to the snapshot
    corrections = []
    for row_id, snap_row in snapshot_rows_by_id.items():
        if row_id not in uploaded_rows_by_id:
            continue  # this row was deleted, already captured in deleted_row_ids
        uploaded_row = uploaded_rows_by_id[row_id]
        for snapshot_col, uploaded_col in diffable_columns.items():
            original_value = snap_row.get(snapshot_col)
            uploaded_value = uploaded_row.get(uploaded_col)
            if values_differ(original_value, uploaded_value):
                corrections.append({
                    "row_index": row_id,
                    "column": snapshot_col,
                    "original_value": "" if original_value is None else str(original_value),
                    "corrected_value": "" if uploaded_value is None else str(uploaded_value),
                })

    return {
        "corrections": corrections,
        "deleted_row_ids": deleted_row_ids,
        "deleted_columns": deleted_columns,
        "renamed_columns": renamed_columns,
        "unresolved_columns": sorted(unresolved_columns),
        "ambiguous_changes": ambiguous_changes,
        "new_row_ids": new_row_ids,
    }